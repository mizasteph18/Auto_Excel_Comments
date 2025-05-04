import openpyxl

def update_excel_with_comments_and_flags(
    main_file_path,
    key_file_path,
    output_file_path,
    header_start_row,
    main_key_column_name,
    main_comments_column_name,
    main_flag_column_name,
    key_sheet_name,
    key_code_column_name,
    key_comment_column_name,
    key_flag_column_name
):
    """
    Updates specified columns ('comments' and 'flag') in a main Excel file
    based on a key Excel file, for sheets where the header row (starting at
    header_start_row) begins with 'comments' (case-insensitive).

    Args:
        main_file_path (str): Path to the main Excel file to be updated.
        key_file_path (str): Path to the key Excel file containing the mapping.
        output_file_path (str): Path to save the updated main Excel file.
        header_start_row (int): Row number (1-based) where the header starts in the main file.
        main_key_column_name (str): Name of the key column in the main file.
        main_comments_column_name (str): Name of the comments column in the main file to update.
        main_flag_column_name (str): Name of the flag column in the main file to update.
        key_sheet_name (str): Name of the sheet in the key file containing the mapping.
        key_code_column_name (str): Name of the code column in the key file.
        key_comment_column_name (str): Name of the comment column in the key file.
        key_flag_column_name (str): Name of the flag column in the key file.
    """
    try:
        # --- Load both Excel workbooks ---
        main_workbook = openpyxl.load_workbook(main_file_path)
        key_workbook = openpyxl.load_workbook(key_file_path)

        # --- Read the key data from the specified sheet ---
        key_data = {}
        if key_sheet_name in key_workbook.sheetnames:
            key_sheet = key_workbook[key_sheet_name]
            key_header_row = [cell.value for cell in key_sheet[1]]  # Assuming header is in the first row of the key sheet
            try:
                key_code_col_index = key_header_row.index(key_code_column_name)
                key_comment_col_index = key_header_row.index(key_comment_column_name)
                key_flag_col_index = key_header_row.index(key_flag_column_name)
                for row in key_sheet.iter_rows(min_row=2):  # Start from the second row after the header
                    code = row[key_code_col_index].value
                    comment = row[key_comment_col_index].value
                    flag = row[key_flag_col_index].value
                    if code is not None:
                        key_data[code] = {'comments': comment, 'flag': flag}
            except ValueError as e:
                print(f"Error: One of the key header names ('{key_code_column_name}', '{key_comment_column_name}', '{key_flag_column_name}') not found in the '{key_sheet_name}' sheet of the key file: {e}")
                return
        else:
            print(f"Error: The sheet named '{key_sheet_name}' was not found in the key file.")
            return

        # --- Iterate through each sheet in the main workbook ---
        for sheet_name in main_workbook.sheetnames:
            main_sheet = main_workbook[sheet_name]

            # --- Check if the header row starts with "comments" (case-insensitive) ---
            header_row = [cell.value for cell in main_sheet[header_start_row]]
            if header_row and header_row[0] is not None and str(header_row[0]).lower().startswith("comments"):
                print(f"Processing sheet: {sheet_name}")

                # --- Identify the column indices in the main file for this sheet ---
                try:
                    main_key_col_index = header_row.index(main_key_column_name)
                    main_comments_col_index = header_row.index(main_comments_column_name)
                    main_flag_col_index = header_row.index(main_flag_column_name)
                except ValueError as e:
                    print(f"Error: One or more of the required main header names ('{main_key_column_name}', '{main_comments_column_name}', '{main_flag_column_name}') not found in sheet '{sheet_name}' on row {header_start_row}: {e}")
                    continue  # Move to the next sheet

                # --- Iterate and update rows in the current worksheet ---
                for row_index in range(header_start_row + 1, main_sheet.max_row + 1):
                    key_value_cell = main_sheet.cell(row=row_index, column=main_key_col_index + 1)
                    key_value = key_value_cell.value

                    if key_value in key_data:
                        comments_value = key_data[key_value]['comments']
                        flag_value = key_data[key_value]['flag']

                        comments_cell = main_sheet.cell(row=row_index, column=main_comments_col_index + 1)
                        flag_cell = main_sheet.cell(row=row_index, column=main_flag_col_index + 1)

                        comments_cell.value = comments_value
                        flag_cell.value = flag_value
            else:
                print(f"Skipping sheet: {sheet_name} - Header on row {header_start_row} does not start with 'comments'.")

        # --- Save the changes to the main Excel file ---
        main_workbook.save(output_file_path)
        print(f"Successfully processed sheets in '{main_file_path}' where the header on row {header_start_row} starts with 'comments' and saved the result to '{output_file_path}'.")

    except FileNotFoundError:
        print("Error: One or both of the specified files were not found.")
    except ValueError as ve:
        print(f"ValueError: {ve}")
    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        # Ensure workbooks are closed
        if 'main_workbook' in locals():
            main_workbook.close()
        if 'key_workbook' in locals():
            key_workbook.close()

if __name__ == "__main__":
    # --- Configuration (Example Usage) ---
    main_file = 'your_main_file.xlsx'
    key_file = 'your_key_file.xlsx'
    output_file = 'updated_main_file.xlsx'
    header_row = 4
    key_column = 'master code'
    comments_column = 'annotation'
    flag_column = 'status'
    key_sheet = 'AutoComment'
    code_key = 'code'
    comment_key = 'comment'
    flag_key = 'flag'

    # --- Call the function ---
    update_excel_with_comments_and_flags(
        main_file_path=main_file,
        key_file_path=key_file,
        output_file_path=output_file,
        header_start_row=header_row,
        main_key_column_name=key_column,
        main_comments_column_name=comments_column,
        main_flag_column_name=flag_column,
        key_sheet_name=key_sheet,
        key_code_column_name=code_key,
        key_comment_column_name=comment_key,
        key_flag_column_name=flag_key
    )
